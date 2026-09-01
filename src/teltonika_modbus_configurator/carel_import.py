"""Carel cDesign Modbus export inspection helpers.

The file loaders are format-independent and the table parser is profile-driven.
Carel remains the first built-in profile while the public Carel API stays
compatible with v0.4 conversion/GUI code.
"""

from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path
import re

from .import_profiles import CAREL_CDESIGN, ImportProfile, carel_name


@dataclass(slots=True)
class CarelImportRow:
    sheet: str
    row_number: int
    name: str = ""
    register: str = ""
    modbus_type: str = ""
    size: str = ""
    data_type: str = ""
    access: str = ""
    raw: tuple[str, ...] = ()


@dataclass(slots=True)
class CarelImportPreview:
    path: str
    sheets: list[str]
    header_row: int | None
    headers: list[str]
    rows: list[CarelImportRow]
    profile_key: str = CAREL_CDESIGN.key
    profile_label: str = CAREL_CDESIGN.label


def sanitize_carel_name(value: str) -> str:
    """Backward-compatible Carel name sanitizer."""
    return carel_name(value)


def _text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _norm(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", value.lower()).strip()


def _column_for(headers: list[str], aliases: tuple[str, ...], *, allow_contains: bool = True) -> int | None:
    normalized = [_norm(h) for h in headers]
    normalized_aliases = [_norm(alias) for alias in aliases]
    for alias in normalized_aliases:
        for i, header in enumerate(normalized):
            if header == alias:
                return i
    if allow_contains:
        for alias in normalized_aliases:
            if len(alias) < 4:
                continue
            for i, header in enumerate(normalized):
                if alias in header:
                    return i
    return None


def detect_header_row(
    rows: list[list[object]],
    max_scan: int = 60,
    *,
    profile: ImportProfile = CAREL_CDESIGN,
) -> int | None:
    """Return the most plausible register-table header for an import profile."""
    best: tuple[int, int] | None = None
    for idx, row in enumerate(rows[:max_scan]):
        headers = [_text(v) for v in row]
        score = 0
        if _column_for(headers, profile.name_aliases) is not None:
            score += 3
        if _column_for(headers, profile.register_aliases) is not None:
            score += 3
        if _column_for(headers, profile.modbus_type_aliases) is not None:
            score += 2
        if _column_for(headers, profile.data_type_aliases) is not None:
            score += 1
        if _column_for(headers, profile.access_aliases) is not None:
            score += 1
        if sum(bool(h) for h in headers) >= 3:
            score += 1
        if score and (best is None or score > best[0]):
            best = (score, idx)
    return None if best is None or best[0] < 4 else best[1]


def preview_rows(
    sheet_name: str,
    rows: list[list[object]],
    *,
    profile: ImportProfile = CAREL_CDESIGN,
) -> tuple[int | None, list[str], list[CarelImportRow]]:
    """Normalize a register table using the selected vendor profile."""
    header_idx = detect_header_row(rows, profile=profile)
    if header_idx is None:
        return None, [], []

    headers = [_text(v) for v in rows[header_idx]]
    name_col = _column_for(headers, profile.name_aliases)
    reg_col = _column_for(headers, profile.register_aliases)
    modbus_type_col = _column_for(headers, profile.modbus_type_aliases)
    size_col = _column_for(headers, profile.size_aliases)
    data_type_col = _column_for(headers, profile.data_type_aliases)
    access_col = _column_for(headers, profile.access_aliases)

    result: list[CarelImportRow] = []
    for source_row, raw_row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
        raw = tuple(_text(v) for v in raw_row)
        if not any(raw):
            continue

        def val(index: int | None) -> str:
            return raw[index] if index is not None and index < len(raw) else ""

        name = profile.sanitize_name(val(name_col))
        register = val(reg_col)
        if not name and not register:
            continue
        result.append(
            CarelImportRow(
                sheet=sheet_name,
                row_number=source_row,
                name=name,
                register=register,
                modbus_type=val(modbus_type_col),
                size=val(size_col),
                data_type=val(data_type_col),
                access=val(access_col),
                raw=raw,
            )
        )
    return header_idx + 1, headers, result


def _preview_from_sheets(
    path: str | Path,
    sheets: list[tuple[str, list[list[object]]]],
    *,
    profile: ImportProfile = CAREL_CDESIGN,
) -> CarelImportPreview:
    best = None
    sheet_names = [name for name, _ in sheets]
    for sheet_name, rows in sheets:
        header_row, headers, parsed = preview_rows(sheet_name, rows, profile=profile)
        candidate = (len(parsed), header_row, headers, parsed)
        if best is None or candidate[0] > best[0]:
            best = candidate
    if best is None:
        return CarelImportPreview(
            str(path), sheet_names, None, [], [],
            profile_key=profile.key, profile_label=profile.label,
        )
    _, header_row, headers, parsed = best
    return CarelImportPreview(
        str(path), sheet_names, header_row, headers, parsed,
        profile_key=profile.key, profile_label=profile.label,
    )


def load_carel_xls(path: str | Path, *, profile: ImportProfile = CAREL_CDESIGN) -> CarelImportPreview:
    """Read a legacy `.xls` export and parse it using a profile."""
    try:
        import xlrd
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError(".xls import requires the 'xlrd' package.") from exc

    workbook = xlrd.open_workbook(str(path), on_demand=True)
    sheets = []
    for sheet_name in workbook.sheet_names():
        sheet = workbook.sheet_by_name(sheet_name)
        sheets.append((sheet_name, [sheet.row_values(i) for i in range(sheet.nrows)]))
    return _preview_from_sheets(path, sheets, profile=profile)


def load_carel_xlsx(path: str | Path, *, profile: ImportProfile = CAREL_CDESIGN) -> CarelImportPreview:
    """Read a modern `.xlsx` export and parse it using a profile."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError(".xlsx import requires the 'openpyxl' package.") from exc

    workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
    sheets = []
    for sheet in workbook.worksheets:
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]
        sheets.append((sheet.title, rows))
    workbook.close()
    return _preview_from_sheets(path, sheets, profile=profile)


def load_carel_csv(path: str | Path, *, profile: ImportProfile = CAREL_CDESIGN) -> CarelImportPreview:
    """Read a CSV export with automatic delimiter sniffing and a profile."""
    text = Path(path).read_text(encoding="utf-8-sig")
    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    rows = [list(row) for row in csv.reader(text.splitlines(), dialect)]
    return _preview_from_sheets(path, [("CSV", rows)], profile=profile)


def load_carel_file(path: str | Path, *, profile: ImportProfile = CAREL_CDESIGN) -> CarelImportPreview:
    """Load XLS/XLSX/CSV input and apply the selected import profile."""
    suffix = Path(path).suffix.lower()
    if suffix == ".xls":
        return load_carel_xls(path, profile=profile)
    if suffix == ".xlsx":
        return load_carel_xlsx(path, profile=profile)
    if suffix == ".csv":
        return load_carel_csv(path, profile=profile)
    raise ValueError(f"Unsupported import format {suffix or '<none>'}. Use .xls, .xlsx, or .csv.")
